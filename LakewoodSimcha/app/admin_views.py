"""
Definition of admin views.
"""

from django.shortcuts import render
from django.shortcuts import redirect
from django.http import HttpRequest
from django.template import RequestContext
from datetime import datetime
from app.models import Event, Customer, Venue
from app.forms import EventForm, CustomerForm, AdminEventForm, UploadFileForm
from django.core.serializers import serialize
from array import array
import json
from django.template.loader import render_to_string
from django.http import JsonResponse
from django.core.mail import send_mail
from django.db.models import F
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from openpyxl import load_workbook
from django.utils import timezone
from django.views.decorators.http import require_http_methods
import pytz
from django.core.exceptions import ObjectDoesNotExist


def event_delete(request, pk):
    event = get_object_or_404(Event, pk=pk)
    data = dict()
    if request.method == 'POST':
        if event.confirmed == 0:
            email_deleted_event(pk)
        if event.customer is not None:
            event.customer.delete()
        event.delete()
        data['form_is_valid'] = True  # This is just to play along with the existing code
        events = Event.objects.filter(venue__id = event.venue.id).filter(start__gte=timezone.now()).all()
        data['html_event_list'] = render_to_string('app/admin/partial_event_list.html', {
            'events': events
        })
    else:
        context = {'event': event}
        data['html_form'] = render_to_string('app/admin/partial_event_delete.html',
            context,
            request=request,
        )
    return JsonResponse(data)

def event_update(request, pk):
    event = get_object_or_404(Event, pk=pk)
    if request.method == 'POST':
        #form = EventForm(request.POST, instance=event)
        form = AdminEventForm(request.POST, instance=event,auto_id=False)
    else:
        #form = EventForm(event.event_type, instance=event)
        form = AdminEventForm(instance=event)
    return save_event_form(request, form, 'app/admin/partial_event_update.html',pk)

def save_event_form(request, form, template_name, pk):
    data = dict()
    if request.method == 'POST':
        if form.is_valid():
            #Leaving this for now as a manual process - venue must delete existing bookings ot let the job expire them
            ##if booking is confirmed, delete all other pending bookings with notifications
            #if form.cleaned_data['confirmed']:
            #    confirm_booking(pk)
            form.save()
            data['form_is_valid'] = True
            #retrieve events for refresh of events list
            events = Event.objects.filter(venue = request.POST['venue']).filter(start__gte=timezone.now()).all()
            data['html_event_list'] = render_to_string('app/admin/partial_event_list.html', {
                'events': events
            })
        else:
            data['form_is_valid'] = False
    context = {'form': form}
    data['html_form'] = render_to_string(template_name, context, request=request)
    return JsonResponse(data)

@login_required(login_url='/login/')
def event_list(request):
    #verify user id
    try:
        if request.user.venue is None:
            return redirect('/login/?next=%s' % request.path)
    except ObjectDoesNotExist:
        return redirect('/login/?next=%s' % request.path)
        
    events = Event.objects.filter(venue__id = request.user.venue.id).filter(start__gte=timezone.now()).all()
    uploadForm = UploadFileForm(initial={'venue':request.user.venue.id})
    return render(request, 'app/admin/event_list.html', {'events': events, 'year':datetime.now().year,'uploadForm':uploadForm})

def email_deleted_event(pk):
    """send an email to customer with an unconfirmed booking that's being deleted."""
    event = Event.objects.get(id=pk)
    try:
        send_mail(
            event.title + ' booking at ' + event.venue.name + ' has been deleted.',
            'Your booking for ' + normalize(event.start).strftime("%A, %d %B %Y %I:%M%p") + ' at ' + event.venue.name + ' has been deleted.\n\nPlease book a new date at ' + event.venue.name,
            'benzyp@yahoo.com',
            [event.customer.email],
            fail_silently=False,
        )
    except Exception as e:
        ex = e

@require_http_methods(["POST"])
def upload_events(request):
    file = request.FILES['file']
    fs = FileSystemStorage()
    venue = request.POST.get('venue')
    #save the file to disk for processing
    filename = fs.save(venue+'/'+file.name, file)
    #BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    media_folder = settings.MEDIA_ROOT
    wb = load_workbook(media_folder + '/'+filename, read_only=True)#open the file and the events worksheet
    ws = wb['Events']
    event_type_to_index = dict()
    for i,j in enumerate(Event.EVENT_TYPES):#load event_type_to_index to get the numeric value of the Event Type
        event_type_to_index.update({j[1]:i})
    for row in ws.iter_rows(row_offset=1):#process the events
        event = Event(title=row[0].value,event_type=(event_type_to_index[row[1].value]+1),description=row[2].value,confirmed=(1 if row[3].value == 'Y' else 0),start=localize(row[4].value),venue_id=venue,created=timezone.now())
        if row[5].value is not None and row[7] is not None:#only save customer info for Tentative bookings - this may change
            customer = Customer(name=row[5].value,email=row[6].value,phone=row[7].value)
            customer.save()
            event.customer = customer
        #only add an event if it doesn't already exist
        event_validate = Event.objects.filter(title=row[0].value).filter(venue_id=venue).filter(start__date=row[4].value.date()).first()
        if event_validate is None:
            event.save()
    events = Event.objects.filter(venue__id = venue).filter(start__gte=timezone.now()).all()
    uploadForm = UploadFileForm(initial={'venue':venue})
    return render(request, 'app/admin/event_list.html', {'events': events, 'year':datetime.now().year,'uploadForm':uploadForm})

def localize(to_convert):
    """Converts naive datetime to localize."""
    tz = pytz.timezone(settings.TIME_ZONE)
    return tz.localize(to_convert)#works on excel value which comes in as dattime

def normalize(to_convert):
    """Converts UTC aware internal times to eastern."""
    tz = pytz.timezone(settings.TIME_ZONE)
    return tz.normalize(to_convert)

#def confirm_booking(pk):
#    event = Event.objects.filter(id=pk).get()
#    #get any other tentative bookings for the same date/venue
#    tentativeEvents = Event.objects.filter(venue__id = event.venue.id).filter(start__gte=event.start.date()).filter(confirmed=0)..all()
#    #for BF and LT allow for double booking on Sunday
#    if event.StopAsyncIteration.weekday() == 6:
#        if event.venue.id == 3:
#            return
#    foreach event in tentativeEvents:
        
#        email_customer()
#        delete
#def upload_events
    